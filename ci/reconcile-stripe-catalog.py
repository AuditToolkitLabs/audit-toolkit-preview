#!/usr/bin/env python3
"""Reconcile a product/payment-link spreadsheet against Stripe objects.

This script validates:
- product IDs exist and names match
- payment link URLs exist and are active
- payment link line items point to the expected product
- price amounts and billing durations align

It uses Stripe CLI auth/context (`stripe whoami`) and live Stripe API calls
through `stripe <resource> <operation>` commands.
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any


HEADER_PRODUCT_FAMILY = "Product "
HEADER_PRODUCT_NAME = "Prodcut Name "
HEADER_PRODUCT_ID = "Prodcut ID"
HEADER_LINK_NAME = "Payment Link Name "
HEADER_LINK_URL = "Related URL "
HEADER_PRICE = "Price "
HEADER_DURATION = "Duration "


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Validate spreadsheet product/payment-link rows against Stripe."
        ),
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        help="Absolute path to the .xlsx catalog file.",
    )
    parser.add_argument(
        "--sheet",
        default="Sheet1",
        help="Sheet name to read (default: Sheet1).",
    )
    parser.add_argument(
        "--mode",
        choices=["live", "test"],
        default="live",
        help="Stripe mode to validate against (default: live).",
    )
    parser.add_argument(
        "--family",
        action="append",
        default=[],
        help=(
            "Optional product family filter, repeatable; matches "
            "the 'Product ' "
            "column case-insensitively."
        ),
    )
    parser.add_argument(
        "--json-out",
        help="Optional file path to write full JSON report.",
    )
    parser.add_argument(
        "--strict-link-name",
        action="store_true",
        help="Also require payment link name to match product name exactly.",
    )
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKC", text)
    replacements = {
        "\u2014": "-",
        "\u2013": "-",
        "\u2212": "-",
        "â€”": "-",
        "â€“": "-",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = " ".join(text.split())
    return text.casefold()


def parse_duration(duration: str) -> list[tuple[str, int]]:
    d = normalize_text(duration)
    if not d:
        return []
    if "year" in d:
        return [("year", 1)]
    if "90" in d and "day" in d:
        return [("day", 90), ("month", 3)]
    if "30" in d and "day" in d:
        return [("day", 30), ("month", 1)]
    if "month" in d:
        return [("month", 1)]
    return []


def amount_to_minor_units(amount: Any) -> int | None:
    if amount is None:
        return None
    try:
        return int(round(float(amount) * 100))
    except (TypeError, ValueError):
        return None


def run_stripe(args: list[str], mode: str) -> dict[str, Any]:
    cmd = ["stripe", *args, f"--{mode}"]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        msg = proc.stderr.strip() or proc.stdout.strip() or "unknown error"
        raise RuntimeError(f"{' '.join(cmd)} failed: {msg}")
    return json.loads(proc.stdout)


def ensure_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency: openpyxl. Install with: pip install openpyxl",
        ) from exc
    return openpyxl


def read_rows(xlsx: Path, sheet: str) -> list[dict[str, Any]]:
    openpyxl = ensure_openpyxl()
    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    if sheet not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise RuntimeError(
            f"Sheet '{sheet}' not found. Available: {available}",
        )
    ws = wb[sheet]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    required = {
        HEADER_PRODUCT_FAMILY,
        HEADER_PRODUCT_NAME,
        HEADER_PRODUCT_ID,
        HEADER_LINK_NAME,
        HEADER_LINK_URL,
        HEADER_PRICE,
        HEADER_DURATION,
    }
    missing = [h for h in required if h not in headers]
    if missing:
        raise RuntimeError(f"Missing expected columns: {', '.join(missing)}")

    records: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None and str(v).strip() for v in values):
            continue
        row = dict(zip(headers, values))
        row["_line"] = r
        records.append(row)
    return records


def list_payment_links(mode: str) -> dict[str, dict[str, Any]]:
    links_by_url: dict[str, dict[str, Any]] = {}
    starting_after: str | None = None

    while True:
        args = ["payment_links", "list", "--limit", "100"]
        if starting_after:
            args.extend(["--starting-after", starting_after])
        page = run_stripe(args, mode=mode)
        data = page.get("data", [])
        for item in data:
            url = item.get("url")
            if url:
                links_by_url[url] = item
        if not page.get("has_more"):
            break
        if not data:
            break
        starting_after = data[-1].get("id")

    return links_by_url


@dataclass
class Caches:
    products: dict[str, dict[str, Any]]
    prices: dict[str, dict[str, Any]]
    payment_links: dict[str, dict[str, Any]]


def validate_row(
    row: dict[str, Any],
    mode: str,
    links_by_url: dict[str, dict[str, Any]],
    caches: Caches,
    strict_link_name: bool,
) -> list[str]:
    issues: list[str] = []

    product_name = str(row.get(HEADER_PRODUCT_NAME, "")).strip()
    product_id = str(row.get(HEADER_PRODUCT_ID, "")).strip()
    link_name = str(row.get(HEADER_LINK_NAME, "")).strip()
    link_url = str(row.get(HEADER_LINK_URL, "")).strip()
    expected_minor = amount_to_minor_units(row.get(HEADER_PRICE))
    allowed_durations = parse_duration(str(row.get(HEADER_DURATION, "")))

    if (
        strict_link_name
        and normalize_text(link_name) != normalize_text(product_name)
    ):
        issues.append(
            "payment-link name mismatch: "
            f"sheet_link='{link_name}' "
            f"sheet_product='{product_name}'",
        )

    try:
        if product_id not in caches.products:
            caches.products[product_id] = run_stripe(
                ["products", "retrieve", product_id],
                mode=mode,
            )
        product = caches.products[product_id]
    except Exception as exc:
        issues.append(f"product lookup failed for {product_id}: {exc}")
        return issues

    if not product.get("active", False):
        issues.append("product is not active")

    stripe_name = str(product.get("name", "")).strip()
    if normalize_text(stripe_name) != normalize_text(product_name):
        issues.append(
            "product name mismatch: "
            f"sheet='{product_name}' stripe='{stripe_name}'",
        )

    payment_link = links_by_url.get(link_url)
    if not payment_link:
        issues.append("payment-link URL not found in Stripe")
        return issues

    if not payment_link.get("active", False):
        issues.append("payment-link is not active")

    pl_id = payment_link.get("id")
    if not pl_id:
        issues.append("payment-link missing id")
        return issues

    try:
        if pl_id not in caches.payment_links:
            caches.payment_links[pl_id] = run_stripe(
                ["payment_links", "retrieve", pl_id, "--expand", "line_items"],
                mode=mode,
            )
        pl_full = caches.payment_links[pl_id]
    except Exception as exc:
        issues.append(f"payment-link retrieve failed for {pl_id}: {exc}")
        return issues

    line_items = ((pl_full.get("line_items") or {}).get("data") or [])
    subscription_data = pl_full.get("subscription_data") or {}
    trial_days = subscription_data.get("trial_period_days")
    if not line_items:
        issues.append("payment-link has no line items")
        return issues

    if len(line_items) != 1:
        issues.append(f"expected 1 line item, found {len(line_items)}")

    matched_product = False
    for line_item in line_items:
        price_obj = line_item.get("price") or {}
        item_product = str(price_obj.get("product") or "")
        if item_product != product_id:
            continue
        matched_product = True

        price_id = str(price_obj.get("id") or "")
        if not price_id:
            issues.append("line-item price id missing")
            continue

        if price_id not in caches.prices:
            caches.prices[price_id] = run_stripe(
                ["prices", "retrieve", price_id],
                mode=mode,
            )
        price = caches.prices[price_id]

        stripe_minor = price.get("unit_amount")
        if expected_minor is not None and stripe_minor != expected_minor:
            issues.append(
                "price mismatch: "
                f"sheet_minor={expected_minor} stripe_minor={stripe_minor}",
            )

        recurring = price.get("recurring") or {}
        interval = recurring.get("interval")
        interval_count = recurring.get("interval_count")
        if allowed_durations:
            if interval and interval_count:
                if (interval, interval_count) not in allowed_durations:
                    issues.append(
                        "duration mismatch: "
                        f"sheet='{row.get(HEADER_DURATION)}' "
                        f"stripe={interval_count} {interval}",
                    )
            elif trial_days:
                trial_duration = ("day", int(trial_days))
                if trial_duration not in allowed_durations:
                    issues.append(
                        "duration mismatch: "
                        f"sheet='{row.get(HEADER_DURATION)}' "
                        f"stripe_trial_days={trial_days}",
                    )
            else:
                issues.append(
                    "duration mismatch: "
                    f"sheet='{row.get(HEADER_DURATION)}' "
                    "stripe has no recurring/trial duration",
                )

        break

    if not matched_product:
        issues.append(
            "payment-link line items do not include expected product id "
            f"'{product_id}'",
        )

    return issues


def main() -> int:
    args = parse_args()
    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(f"ERROR: xlsx file not found: {xlsx}", file=sys.stderr)
        return 2

    try:
        records = read_rows(xlsx, args.sheet)
    except Exception as exc:
        print(f"ERROR: could not read workbook: {exc}", file=sys.stderr)
        return 2

    if args.family:
        selected = {normalize_text(v) for v in args.family}
        records = [
            r
            for r in records
            if normalize_text(r.get(HEADER_PRODUCT_FAMILY, "")) in selected
        ]

    try:
        links_by_url = list_payment_links(mode=args.mode)
    except Exception as exc:
        print(
            f"ERROR: failed listing Stripe payment links: {exc}",
            file=sys.stderr,
        )
        return 2

    caches = Caches(products={}, prices={}, payment_links={})

    results: list[dict[str, Any]] = []
    mismatches = 0
    for row in records:
        issues = validate_row(
            row=row,
            mode=args.mode,
            links_by_url=links_by_url,
            caches=caches,
            strict_link_name=args.strict_link_name,
        )
        if issues:
            mismatches += 1
        results.append(
            {
                "line": row["_line"],
                "family": row.get(HEADER_PRODUCT_FAMILY),
                "product_name": row.get(HEADER_PRODUCT_NAME),
                "product_id": row.get(HEADER_PRODUCT_ID),
                "payment_link_name": row.get(HEADER_LINK_NAME),
                "payment_link_url": row.get(HEADER_LINK_URL),
                "price": row.get(HEADER_PRICE),
                "duration": row.get(HEADER_DURATION),
                "issues": issues,
            },
        )

    report = {
        "mode": args.mode,
        "xlsx": str(xlsx),
        "sheet": args.sheet,
        "rows_scanned": len(records),
        "rows_with_issues": mismatches,
        "rows_ok": len(records) - mismatches,
        "results": results,
    }

    print(f"rows_scanned={report['rows_scanned']}")
    print(f"rows_ok={report['rows_ok']}")
    print(f"rows_with_issues={report['rows_with_issues']}")
    for item in results:
        if not item["issues"]:
            continue
        print(f"line {item['line']} :: {item['product_name']}")
        for issue in item["issues"]:
            print(f"  - {issue}")

    if args.json_out:
        out_path = Path(args.json_out)
        out_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
        print(f"json_report={out_path}")

    return 1 if mismatches else 0


if __name__ == "__main__":
    raise SystemExit(main())
